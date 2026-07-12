"""Reads any Excel file entirely according to a YAML mapping.

Per Iteration 3 Task A, this is the single reader for every input
slot — it replaces the Iteration 1/2 split between `BaseFileReader`
(hard-coded "the system file is special") and `SecondaryFileReader`
(already mapping-driven). No input file is more special than another
at the code level: the "system file" is simply whichever mapping the
caller assigns to a slot.

Uses `read_only=True` and only reads the specific columns the mapping
declares (rather than every column up to the sheet's width), since
real-world exports often have dozens of unused columns per row.
"""

from dataclasses import replace
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

from src.application.dto.parcel_record import ParcelRecord
from src.application.ports.column_mapper_port import ColumnMapperPort
from src.application.ports.data_source_port import DataSourcePort
from src.infrastructure.excel.cell_parsing import clean_text
from src.shared.arabic_normalizer import normalize_arabic


class MappedFileReader(DataSourcePort):
    """Reads an Excel file into `ParcelRecord`s using a YAML mapping.

    The join key (`holding_id_raw`) is always extracted directly from
    the mapping's `join_key_column` — independent of the `ColumnMapperPort`,
    which maps every other field. This means the join key never depends
    on a source declaring a semantically-named field for it.
    """

    def __init__(
        self,
        path: Path,
        mapper: ColumnMapperPort,
        config: dict[str, Any],
        apply_exclusion: bool = True,
    ) -> None:
        self._path = path
        self._mapper = mapper
        self._config = config
        self._apply_exclusion = apply_exclusion
        self._excluded_count = 0
        self._sheet_fallback_warning: str | None = None

    @property
    def excluded_count(self) -> int:
        """How many rows the last `read()` call filtered out via the exclusion rule.

        Zero before `read()` has been called, and while `apply_exclusion`
        is False (nothing is excluded in that mode).
        """
        return self._excluded_count

    @property
    def sheet_fallback_warning(self) -> str | None:
        """See `DataSourcePort.sheet_fallback_warning`."""
        return self._sheet_fallback_warning

    def read(self) -> list[ParcelRecord]:
        """Read, exclude, and map all data rows to `ParcelRecord`s."""
        workbook = openpyxl.load_workbook(self._path, data_only=True, read_only=True)
        worksheet = self._resolve_worksheet(workbook)

        data_start = int(self._config["data_starts_at_row"])
        join_key_column: str = self._config["join_key_column"]
        exclude_config = self._config.get("exclude_when") if self._apply_exclusion else None
        fields: dict[str, str] = self._config["fields"]

        needed_columns = set(fields.values()) | {join_key_column}
        if exclude_config:
            needed_columns.add(exclude_config["column"])
        column_indices = {col: column_index_from_string(col) for col in needed_columns}
        max_col = max(column_indices.values())

        self._excluded_count = 0
        records: list[ParcelRecord] = []
        for row in worksheet.iter_rows(min_row=data_start, max_col=max_col):
            raw_row = {col: row[index - 1].value for col, index in column_indices.items()}
            if self._is_excluded(raw_row, exclude_config):
                self._excluded_count += 1
                continue

            holding_id = clean_text(raw_row.get(join_key_column))
            if holding_id is None:
                continue

            record = self._mapper.map(raw_row)
            records.append(replace(record, holding_id_raw=holding_id))

        return records

    def _resolve_worksheet(self, workbook: Any) -> Any:
        worksheet, self._sheet_fallback_warning = resolve_worksheet(
            workbook, self._config.get("sheet_name"), self._path.name
        )
        return worksheet

    def _is_excluded(self, raw_row: dict[str, Any], exclude_config: dict[str, Any] | None) -> bool:
        if not exclude_config:
            return False
        value = raw_row.get(exclude_config["column"])
        if value is None:
            return False
        normalized_value = normalize_arabic(str(value))
        excluded_values = {normalize_arabic(v) for v in exclude_config["values"]}
        return normalized_value in excluded_values


def resolve_worksheet(workbook: Any, sheet_name: str | None, source_filename: str) -> Any:
    """Resolve `sheet_name` within `workbook`, with a single-sheet fallback.

    Shared by `MappedFileReader` and `read_first_field_value` so both
    apply the same "different export runs sometimes get a slightly
    different auto-generated sheet name" tolerance (see
    `MappedFileReader._resolve_worksheet`'s original docstring).

    Returns:
        A `(worksheet, fallback_warning)` tuple; `fallback_warning` is
        None unless the fallback was used.
    """
    if not sheet_name:
        return workbook.active, None

    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name], None

    if len(workbook.sheetnames) == 1:
        actual_name = workbook.sheetnames[0]
        warning = (
            f"Expected sheet '{sheet_name}' not found in '{source_filename}'; "
            f"used its only sheet '{actual_name}' instead."
        )
        return workbook[actual_name], warning

    raise KeyError(
        f"Worksheet '{sheet_name}' not found in '{source_filename}', and the "
        f"workbook has multiple sheets ({workbook.sheetnames}), so the correct "
        f"one can't be inferred."
    )


def read_first_field_value(
    path: Path, config: dict[str, Any], field_name: str, max_rows_to_scan: int = 50
) -> str | None:
    """Peek the first non-empty value of one mapped field, without a full read.

    Used where only a single representative value is needed (e.g. the
    society name for the output filename) rather than a full
    per-row `ParcelRecord` pass. Scans at most `max_rows_to_scan` data
    rows in case the first few are blank.

    Returns:
        The cleaned text value, or None if `field_name` isn't declared
        in `config["fields"]` or every scanned row is blank there.
    """
    column = config["fields"].get(field_name)
    if not column:
        return None

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    worksheet, _ = resolve_worksheet(workbook, config.get("sheet_name"), path.name)

    data_start = int(config["data_starts_at_row"])
    col_index = column_index_from_string(column)
    for row in worksheet.iter_rows(
        min_row=data_start, max_row=data_start + max_rows_to_scan - 1, max_col=col_index
    ):
        value = clean_text(row[col_index - 1].value)
        if value is not None:
            return value
    return None
