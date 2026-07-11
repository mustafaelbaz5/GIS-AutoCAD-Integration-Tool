"""Reads a previously generated output file back into memory.

Per Iteration 2 Task C (Search feature) and §8.2's error-handling rule:
a malformed file (missing expected columns) raises `OutputFileFormatError`
naming exactly which columns are missing, so the caller can show a
precise Arabic message rather than a raw traceback.
"""

from pathlib import Path
from typing import Any

import openpyxl

from src.application.ports.output_file_source_port import OutputFileSourcePort
from src.domain.entities.basin import Basin
from src.domain.entities.borders import Borders
from src.domain.entities.holder import Holder
from src.domain.entities.parcel import Parcel
from src.domain.value_objects.area import Area
from src.domain.value_objects.holding_id import HoldingId
from src.infrastructure.excel.output_schema import OUTPUT_COLUMNS

_TOTALS_LABEL = "الإجمالي"
_HOLDING_ID_HEADER = "رقم الحيازة"


class OutputFileFormatError(Exception):
    """Raised when an output file is missing one or more expected columns."""

    def __init__(self, missing_headers: list[str]) -> None:
        self.missing_headers = missing_headers
        super().__init__(f"Missing expected columns: {', '.join(missing_headers)}")


class OutputFileReader(OutputFileSourcePort):
    """Reads a previously generated output file into `Parcel` entities.

    Locates columns by header text (not fixed position), so it still
    works if columns were reordered, and can precisely report which
    expected headers are missing rather than failing opaquely.
    """

    def read(self, path: Path) -> list[Parcel]:
        """Read all parcels from the output file at `path`.

        Raises:
            OutputFileFormatError: If any expected output column header
                is missing from the file's header row.
        """
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        worksheet = workbook.active

        header_to_col = self._map_headers(worksheet)
        self._validate_headers(header_to_col)

        parcels: list[Parcel] = []
        holding_id_col = header_to_col[_HOLDING_ID_HEADER]
        for row in worksheet.iter_rows(min_row=2):
            holding_id_cell = row[holding_id_col - 1].value
            if holding_id_cell is None:
                continue
            holding_id_text = str(holding_id_cell).strip()
            if not holding_id_text or holding_id_text == _TOTALS_LABEL:
                continue
            parcels.append(self._build_parcel(row, header_to_col, holding_id_text))

        return parcels

    def _map_headers(self, worksheet: Any) -> dict[str, int]:
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
        return {cell.value: cell.column for cell in header_row if cell.value}

    def _validate_headers(self, header_to_col: dict[str, int]) -> None:
        expected = [column.header for column in OUTPUT_COLUMNS]
        missing = [header for header in expected if header not in header_to_col]
        if missing:
            raise OutputFileFormatError(missing)

    def _build_parcel(
        self, row: tuple[Any, ...], header_to_col: dict[str, int], holding_id_text: str
    ) -> Parcel:
        def text(header: str) -> str | None:
            value = row[header_to_col[header] - 1].value
            if value is None:
                return None
            stripped = str(value).strip()
            return stripped or None

        def number(header: str) -> float | None:
            value = row[header_to_col[header] - 1].value
            if value is None:
                return None
            if isinstance(value, int | float):
                return float(value)
            try:
                return float(str(value))
            except ValueError:
                return None

        return Parcel(
            holding_id=HoldingId(holding_id_text),
            page_number=text("رقم الصفحة بالسجل"),
            directorate=text("المديريه"),
            administration=text("الأداره"),
            basin=Basin(name=text("اسم الحوض"), code=text("كود الحوض")),
            holder=Holder(name=text("اسم الحائز"), national_id=text("الرقم القومي")),
            borders=Borders(
                east=text("الحد الشرقى"),
                south=text("الحد القبلى"),
                west=text("الحد الغربى"),
                north=text("الحد البحرى"),
            ),
            land_number=text("رقم الأرض"),
            area=Area(feddan=number("فدان"), qirat=number("قيراط"), sahm=number("سهم")),
        )
