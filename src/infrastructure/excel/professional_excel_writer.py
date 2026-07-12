"""Professional Excel writer for the final merged output, per §8 and
Iteration 2 Task B (typography/column-width refinements)."""

from collections.abc import Sequence
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.application.ports.data_sink_port import DataSinkPort
from src.domain.entities.parcel import Parcel
from src.domain.services.area_calculator import AreaCalculator
from src.infrastructure.excel.formatting_config import DEFAULT_FORMATTING_CONFIG, FormattingConfig
from src.infrastructure.excel.output_schema import OUTPUT_COLUMNS

_HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
_HEADER_TEXT_COLOR = "F1F5F9"
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

_THIN_SIDE = Side(style="thin", color="CBD5E1")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_BAND_FILL_A = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_BAND_FILL_B = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

_NUMBER_FORMAT = "0.00"
_TEXT_FORMAT = "@"
_TOTALS_LABEL = "الإجمالي"
_SHEET_TITLE = "البيانات المجمعة"
_NO_BASIN_LABEL = "بدون حوض"

_NO_BASIN_YET: object = object()


def default_output_filename(now: datetime | None = None) -> str:
    """Return the timestamped default filename, per §8.1."""
    timestamp = (now or datetime.now()).strftime("%Y-%m-%d_%H%M%S")
    return f"merged_output_{timestamp}.xlsx"


_FILENAME_SUFFIX = "كامل"
_ILLEGAL_FILENAME_CHARS = '\\/:*?"<>|'


def build_output_filename(society_name: str | None, now: datetime | None = None) -> str:
    """Return the output filename, derived from the society name when available.

    Per the user-specified convention: takes the part of `society_name`
    before its first "-" (e.g. "ميت فضالة-الائتمان الزراعي" becomes
    "ميت فضالة"), appends the fixed suffix "كامل" and today's date —
    e.g. "ميت فضالة_كامل_2026-07-12.xlsx". Falls back to
    `default_output_filename` when no society name is available (e.g.
    the primary mapping doesn't declare a society column, or every row
    in the file has one blank).
    """
    if society_name:
        prefix = society_name.split("-", 1)[0].strip()
        prefix = "".join(ch for ch in prefix if ch not in _ILLEGAL_FILENAME_CHARS).strip()
        if prefix:
            date_str = (now or datetime.now()).strftime("%Y-%m-%d")
            return f"{prefix}_{_FILENAME_SUFFIX}_{date_str}.xlsx"
    return default_output_filename(now)


def resolve_unique_path(path: Path) -> Path:
    """Return `path` if free, otherwise the first free "(1)", "(2)", ... variant.

    The default filename is timestamped to the second, so two runs
    started in the same second (or a pre-existing file with that exact
    name) would otherwise silently overwrite each other.
    """
    if not path.exists():
        return path
    index = 1
    while True:
        candidate = path.with_name(f"{path.stem}({index}){path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


class ProfessionalExcelWriter(DataSinkPort):
    """Writes the final merged parcel list as a formatted .xlsx file.

    Implements the formatting rules of project brief §8: styled header,
    RTL sheet direction, frozen header row, thin borders, banded fill
    per basin block with a blank separator row between basins, and a
    totals row. Column widths, font sizes/family, and row heights come
    from an injected `FormattingConfig` (Iteration 2 Task B) rather
    than being hard-coded, so they're a one-file tuning knob.
    """

    def __init__(self, config: FormattingConfig = DEFAULT_FORMATTING_CONFIG) -> None:
        self._config = config
        self._header_font = Font(
            name=config.font_family,
            size=config.header_font_size,
            color=_HEADER_TEXT_COLOR,
            bold=True,
        )
        self._body_font = Font(name=config.font_family, size=config.body_font_size)
        self._totals_font = Font(name=config.font_family, size=config.body_font_size, bold=True)

    def write(self, parcels: Sequence[Parcel], output_path: Path) -> None:
        """Write `parcels` to `output_path` as a formatted .xlsx file.

        Also writes one additional sheet per distinct اسم الحوض value,
        each containing only that basin's rows in the same format as
        the main sheet — per the user's request to be able to work
        with each basin's data separately without re-sorting the main
        sheet (which stays basin-block-ordered, as it already was).
        """
        workbook = Workbook()
        main_worksheet = workbook.active
        main_worksheet.title = _SHEET_TITLE
        self._write_full_sheet(main_worksheet, parcels)

        for basin_name, basin_parcels in self._group_by_basin(parcels).items():
            sheet_title = self._unique_sheet_title(workbook, basin_name)
            basin_worksheet = workbook.create_sheet(sheet_title)
            self._write_full_sheet(basin_worksheet, basin_parcels)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

    def _write_full_sheet(self, worksheet: Worksheet, parcels: Sequence[Parcel]) -> None:
        worksheet.sheet_view.rightToLeft = True
        self._write_header(worksheet)
        last_row = self._write_data_rows(worksheet, parcels)
        self._write_totals_row(worksheet, last_row, parcels)
        self._autofit_columns(worksheet, parcels)
        worksheet.freeze_panes = "A2"

    def _group_by_basin(self, parcels: Sequence[Parcel]) -> dict[str, list[Parcel]]:
        """Group `parcels` by اسم الحوض, preserving first-appearance order.

        Parcels with no basin name are grouped under `_NO_BASIN_LABEL`
        rather than dropped, so every row still appears in some basin
        sheet.
        """
        groups: dict[str, list[Parcel]] = {}
        for parcel in parcels:
            key = parcel.basin.name or _NO_BASIN_LABEL
            groups.setdefault(key, []).append(parcel)
        return groups

    def _unique_sheet_title(self, workbook: Workbook, basin_name: str) -> str:
        """Sanitize `basin_name` into a valid, unique Excel sheet title.

        Excel sheet names must be <=31 chars and may not contain
        `: \\ / ? * [ ]`. Truncation could otherwise collide two
        distinct long basin names into the same title, so a numeric
        suffix is appended on collision.
        """
        sanitized = "".join(ch for ch in basin_name if ch not in "\\/*?:[]")
        sanitized = sanitized.strip() or _NO_BASIN_LABEL
        title = sanitized[:31]
        existing = set(workbook.sheetnames)
        if title not in existing:
            return title
        suffix = 2
        while True:
            candidate = f"{sanitized[: 31 - len(str(suffix)) - 1]}_{suffix}"
            if candidate not in existing:
                return candidate
            suffix += 1

    def _write_header(self, worksheet: Worksheet) -> None:
        worksheet.row_dimensions[1].height = self._config.header_row_height
        for col_index, column in enumerate(OUTPUT_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_index, value=column.header)
            cell.fill = _HEADER_FILL
            cell.font = self._header_font
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

    def _write_data_rows(self, worksheet: Worksheet, parcels: Sequence[Parcel]) -> int:
        row_index = 2
        current_basin: object = _NO_BASIN_YET
        band_toggle = False

        for parcel in parcels:
            basin_key = parcel.basin.name
            if basin_key != current_basin:
                if current_basin is not _NO_BASIN_YET:
                    row_index += 1  # blank separator row between basin blocks
                current_basin = basin_key
                band_toggle = not band_toggle
            self._write_row(worksheet, row_index, parcel, band_toggle)
            row_index += 1

        return row_index - 1

    def _write_row(
        self, worksheet: Worksheet, row_index: int, parcel: Parcel, band_toggle: bool
    ) -> None:
        worksheet.row_dimensions[row_index].height = self._config.data_row_height
        fill = _BAND_FILL_B if band_toggle else _BAND_FILL_A
        for col_index, column in enumerate(OUTPUT_COLUMNS, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=column.getter(parcel))
            cell.fill = fill
            cell.font = self._body_font
            cell.border = _THIN_BORDER
            if column.is_numeric:
                cell.number_format = _NUMBER_FORMAT
            elif column.is_text_national_id:
                cell.number_format = _TEXT_FORMAT

    def _write_totals_row(
        self, worksheet: Worksheet, last_data_row: int, parcels: Sequence[Parcel]
    ) -> None:
        row_index = last_data_row + 1
        worksheet.row_dimensions[row_index].height = self._config.data_row_height
        totals = self._compute_totals(parcels)

        label_cell = worksheet.cell(row=row_index, column=1, value=_TOTALS_LABEL)
        label_cell.font = self._totals_font
        label_cell.border = _THIN_BORDER

        for col_index, column in enumerate(OUTPUT_COLUMNS, start=1):
            if column.header not in totals:
                continue
            cell = worksheet.cell(row=row_index, column=col_index, value=totals[column.header])
            cell.font = self._totals_font
            cell.border = _THIN_BORDER
            cell.number_format = _NUMBER_FORMAT

    def _compute_totals(self, parcels: Sequence[Parcel]) -> dict[str, float]:
        feddan_total = sum(p.area.feddan or 0.0 for p in parcels)
        qirat_total = sum(p.area.qirat or 0.0 for p in parcels)
        sahm_total = sum(p.area.sahm or 0.0 for p in parcels)
        sqm_total = sum(AreaCalculator.calculate_total_sqm(p.area) or 0.0 for p in parcels)
        return {
            "فدان": round(feddan_total, 2),
            "قيراط": round(qirat_total, 2),
            "سهم": round(sahm_total, 2),
            "إجمالي المساحة (م²)": round(sqm_total, 2),
        }

    def _autofit_columns(self, worksheet: Worksheet, parcels: Sequence[Parcel]) -> None:
        for col_index, column in enumerate(OUTPUT_COLUMNS, start=1):
            max_length = len(column.header)
            for parcel in parcels:
                value = column.getter(parcel)
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            width = self._config.column_width(column.header, max_length)
            worksheet.column_dimensions[get_column_letter(col_index)].width = width
