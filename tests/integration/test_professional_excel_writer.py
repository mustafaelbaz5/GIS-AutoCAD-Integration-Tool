"""Integration tests for ProfessionalExcelWriter, per project brief §8."""

from datetime import datetime
from pathlib import Path

import openpyxl
from src.domain.entities.basin import Basin
from src.domain.entities.borders import Borders
from src.domain.entities.holder import Holder
from src.domain.entities.parcel import Parcel
from src.domain.value_objects.area import Area
from src.domain.value_objects.holding_id import HoldingId
from src.infrastructure.excel.formatting_config import FormattingConfig
from src.infrastructure.excel.output_schema import OUTPUT_COLUMNS
from src.infrastructure.excel.professional_excel_writer import (
    ProfessionalExcelWriter,
    default_output_filename,
)


def make_parcel(
    holding_id: str,
    basin_name: str | None = "basin",
    holder_name: str | None = "Holder",
    national_id: str | None = None,
    feddan: float | None = 1.0,
    qirat: float | None = 2.0,
    sahm: float | None = 3.0,
) -> Parcel:
    return Parcel(
        holding_id=HoldingId(holding_id),
        page_number="1",
        directorate="Directorate",
        administration="Administration",
        basin=Basin(name=basin_name, code="10"),
        holder=Holder(name=holder_name, national_id=national_id),
        borders=Borders(east="East", south="South", west="West", north="North"),
        land_number="Land1",
        area=Area(feddan=feddan, qirat=qirat, sahm=sahm),
    )


def test_writes_header_row_with_all_columns_in_order(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    ProfessionalExcelWriter().write([make_parcel("1")], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active
    header_values = [
        worksheet.cell(row=1, column=i).value for i in range(1, len(OUTPUT_COLUMNS) + 1)
    ]

    assert header_values == [c.header for c in OUTPUT_COLUMNS]


def test_writes_data_row_values(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    parcel = make_parcel("010", holder_name="Some Holder")
    ProfessionalExcelWriter().write([parcel], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    assert worksheet.cell(row=2, column=1).value == "010"
    assert worksheet.cell(row=2, column=7).value == "Some Holder"


def test_national_id_column_is_text_formatted(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    parcel = make_parcel("1", national_id="01234567890123")
    ProfessionalExcelWriter().write([parcel], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active
    cell = worksheet.cell(row=2, column=8)

    assert cell.value == "01234567890123"
    assert cell.number_format == "@"


def test_area_columns_are_numeric_not_string(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    parcel = make_parcel("1", feddan=2.5, qirat=3.0, sahm=4.0)
    ProfessionalExcelWriter().write([parcel], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    assert isinstance(worksheet.cell(row=2, column=14).value, int | float)
    assert isinstance(worksheet.cell(row=2, column=17).value, int | float)


def test_rtl_direction_applied(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    ProfessionalExcelWriter().write([make_parcel("1")], output_path)

    workbook = openpyxl.load_workbook(output_path)
    assert workbook.active.sheet_view.rightToLeft is True


def test_freeze_panes_applied(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    ProfessionalExcelWriter().write([make_parcel("1")], output_path)

    workbook = openpyxl.load_workbook(output_path)
    assert workbook.active.freeze_panes == "A2"


def test_totals_row_sums_area_columns(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    parcels = [
        make_parcel("1", basin_name="A", feddan=1.0, qirat=2.0, sahm=3.0),
        make_parcel("2", basin_name="A", feddan=4.0, qirat=5.0, sahm=6.0),
    ]
    ProfessionalExcelWriter().write(parcels, output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active
    totals_row = worksheet.max_row

    assert worksheet.cell(row=totals_row, column=1).value == "الإجمالي"
    assert worksheet.cell(row=totals_row, column=14).value == 5.0
    assert worksheet.cell(row=totals_row, column=15).value == 7.0
    assert worksheet.cell(row=totals_row, column=16).value == 9.0


def test_blank_separator_row_between_basin_blocks(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    parcels = [
        make_parcel("1", basin_name="A"),
        make_parcel("2", basin_name="B"),
    ]
    ProfessionalExcelWriter().write(parcels, output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    assert worksheet.cell(row=2, column=1).value == "1"
    assert worksheet.cell(row=3, column=1).value is None
    assert worksheet.cell(row=4, column=1).value == "2"


def test_basin_blocks_use_alternating_fill(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    parcels = [
        make_parcel("1", basin_name="A"),
        make_parcel("2", basin_name="B"),
    ]
    ProfessionalExcelWriter().write(parcels, output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    fill_a = worksheet.cell(row=2, column=1).fill.start_color.rgb
    fill_b = worksheet.cell(row=4, column=1).fill.start_color.rgb
    assert fill_a != fill_b


def test_creates_output_directory_if_missing(tmp_path: Path) -> None:
    output_path = tmp_path / "nested" / "dir" / "output.xlsx"
    ProfessionalExcelWriter().write([make_parcel("1")], output_path)

    assert output_path.exists()


def test_output_file_reopens_without_error(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    ProfessionalExcelWriter().write([make_parcel("1"), make_parcel("2")], output_path)

    workbook = openpyxl.load_workbook(output_path)
    assert workbook.active is not None


def test_default_output_filename_format() -> None:
    filename = default_output_filename(now=datetime(2026, 7, 11, 14, 30, 22))
    assert filename == "merged_output_2026-07-11_143022.xlsx"


def test_header_and_body_font_sizes_per_config(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    ProfessionalExcelWriter().write([make_parcel("1")], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    header_font = worksheet.cell(row=1, column=1).font
    body_font = worksheet.cell(row=2, column=1).font

    assert header_font.size == 14
    assert header_font.bold is True
    assert body_font.size == 13


def test_header_and_data_row_heights_per_config(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    ProfessionalExcelWriter().write([make_parcel("1")], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    assert worksheet.row_dimensions[1].height == 22.0
    assert worksheet.row_dimensions[2].height == 18.0


def test_narrow_columns_cap_at_twelve_even_with_long_content(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    parcel = make_parcel("123456789012345678901234567890")
    ProfessionalExcelWriter().write([parcel], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    holding_id_column_letter = "A"  # رقم الحيازة, first output column
    assert worksheet.column_dimensions[holding_id_column_letter].width == 12


def test_wide_columns_never_exceed_max_width_but_truncate_only_visually(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    long_name = "س" * 100
    parcel = make_parcel("1", holder_name=long_name)
    ProfessionalExcelWriter().write([parcel], output_path)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook.active

    holder_column_letter = "G"  # اسم الحائز
    assert worksheet.column_dimensions[holder_column_letter].width == 35
    # Full value must still be present in the cell, not truncated on disk.
    assert worksheet.cell(row=2, column=7).value == long_name


def test_column_width_is_config_driven_not_hard_coded(tmp_path: Path) -> None:
    """Proves Task B acceptance criterion D: a config change changes the output."""
    default_path = tmp_path / "default.xlsx"
    custom_path = tmp_path / "custom.xlsx"
    long_name = "س" * 100
    parcel = make_parcel("1", holder_name=long_name)

    ProfessionalExcelWriter().write([parcel], default_path)
    ProfessionalExcelWriter(FormattingConfig(max_column_width=20)).write([parcel], custom_path)

    default_width = openpyxl.load_workbook(default_path).active.column_dimensions["G"].width
    custom_width = openpyxl.load_workbook(custom_path).active.column_dimensions["G"].width

    assert default_width == 35
    assert custom_width == 20
    assert default_width != custom_width
