"""Integration tests for MappedFileReader, per Iteration 3 Task A.

Consolidates what were previously separate BaseFileReader and
SecondaryFileReader test suites — both file types now go through the
exact same reader class, driven entirely by their YAML mapping. The
real secondary file contains duplicate رقم حيازه 2022 values across
different village sections, so tests derive expected values directly
from the raw workbook at run time rather than hardcoding a single
row's data (see the Iteration-1 test history for why).
"""

from collections.abc import Iterator
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from src.infrastructure.config.yaml_mapping_loader import load_mapping
from src.infrastructure.excel.mapped_file_reader import MappedFileReader, read_first_field_value
from src.infrastructure.excel.yaml_column_mapper import YamlColumnMapper
from src.shared.arabic_normalizer import normalize_arabic

_MAPPINGS_DIR = (
    Path(__file__).resolve().parents[2] / "src" / "infrastructure" / "config" / "default_mappings"
)
SYSTEM_MAPPING_PATH = _MAPPINGS_DIR / "system_file_default.yaml"
EXTERNAL_MAPPING_PATH = _MAPPINGS_DIR / "external_file_default.yaml"
SEASONAL_MAPPING_PATH = _MAPPINGS_DIR / "seasonal_survey_default.yaml"

SEASONAL_SHEET_NAME = "سجل 2 خدمات"
SEASONAL_DATA_START_ROW = 7
_EXCLUDED_TARGETS = {normalize_arabic("لاغى"), normalize_arabic("لاغي")}


def _read_system_records(base_file_path: Path) -> list[Any]:
    config = load_mapping(SYSTEM_MAPPING_PATH)
    mapper = YamlColumnMapper(config["fields"])
    return MappedFileReader(base_file_path, mapper, config).read()


def _read_external_records(secondary_file_path: Path) -> list[Any]:
    config = load_mapping(EXTERNAL_MAPPING_PATH)
    mapper = YamlColumnMapper(config["fields"])
    return MappedFileReader(secondary_file_path, mapper, config).read()


def _read_seasonal_records(
    seasonal_secondary_file_path: Path, apply_exclusion: bool = True
) -> list[Any]:
    config = load_mapping(SEASONAL_MAPPING_PATH)
    mapper = YamlColumnMapper(config["fields"])
    return MappedFileReader(
        seasonal_secondary_file_path, mapper, config, apply_exclusion=apply_exclusion
    ).read()


def _seasonal_raw_rows(seasonal_secondary_file_path: Path) -> Iterator[dict[str, Any]]:
    workbook = openpyxl.load_workbook(seasonal_secondary_file_path, data_only=True)
    worksheet = workbook[SEASONAL_SHEET_NAME]
    for row_number in range(SEASONAL_DATA_START_ROW, worksheet.max_row + 1):
        yield {
            col: worksheet[f"{col}{row_number}"].value
            for col in ("K", "M", "N", "AA", "AB", "AC", "AD", "AF")
        }


def _is_excluded_row(cells: dict[str, Any]) -> bool:
    m_value = cells["M"]
    return m_value is not None and normalize_arabic(str(m_value)) in _EXCLUDED_TARGETS


# --- System-file mapping (real file, parcel-level "lands report" format) ---


def test_system_mapping_reads_all_1137_data_rows(base_file_path: Path) -> None:
    records = _read_system_records(base_file_path)
    assert len(records) == 1137


def test_system_mapping_preserves_leading_zeros(base_file_path: Path) -> None:
    records = _read_system_records(base_file_path)
    assert records[0].holding_id_raw == "00240"


def test_system_mapping_known_first_row_holder_and_basin(base_file_path: Path) -> None:
    records = _read_system_records(base_file_path)
    first = records[0]
    assert first.basin_name == "الكبيرة الغربى"
    assert first.holder_name == "اسلام محمد محمد سرحان"
    assert first.directorate == "الدقهليه"
    assert first.administration == "اجا"


def test_system_mapping_border_placeholders_are_normalized_to_none(base_file_path: Path) -> None:
    records = _read_system_records(base_file_path)
    for placeholder in ("-", "_"):
        assert all(r.east != placeholder for r in records)
        assert all(r.west != placeholder for r in records)
        assert all(r.south != placeholder for r in records)
        assert all(r.north != placeholder for r in records)


def test_system_mapping_basin_code_column_is_empty(base_file_path: Path) -> None:
    """This report never populates كود الحوض — only اسم الحوض."""
    records = _read_system_records(base_file_path)
    assert all(r.basin_code is None for r in records)


def test_system_mapping_never_fabricates_national_id(base_file_path: Path) -> None:
    """This report has no national-ID column — must stay None, never guessed."""
    records = _read_system_records(base_file_path)
    assert all(r.national_id is None for r in records)


# --- External-file mapping (real file, "approved holdings" format) ---------


def test_external_mapping_reads_all_928_data_rows(secondary_file_path: Path) -> None:
    records = _read_external_records(secondary_file_path)
    assert len(records) == 928


def test_external_mapping_join_key_comes_from_column_r(secondary_file_path: Path) -> None:
    records = _read_external_records(secondary_file_path)
    assert records[0].holding_id_raw == "221"
    assert records[0].holder_name == "ابتسام حسن حسن البرقى"


def test_external_mapping_national_id_is_read_directly(secondary_file_path: Path) -> None:
    records = _read_external_records(secondary_file_path)
    assert records[0].national_id == "26711121202206"
    assert all(r.national_id is None or len(r.national_id) == 14 for r in records)


def test_external_mapping_area_fields_are_read(secondary_file_path: Path) -> None:
    records = _read_external_records(secondary_file_path)
    first = records[0]
    assert first.feddan == 0.0
    assert first.qirat == 13.0
    assert first.sahm == 0.0


def test_external_mapping_never_fabricates_unmapped_fields(secondary_file_path: Path) -> None:
    """This report has no basin/border/land-number columns at all — every
    row must resolve those fields to None, never a guessed value."""
    records = _read_external_records(secondary_file_path)
    assert all(r.basin_name is None for r in records)
    assert all(r.basin_code is None for r in records)
    assert all(r.east is None for r in records)
    assert all(r.land_number is None for r in records)


def test_base_and_external_join_keys_fully_overlap(
    base_file_path: Path, secondary_file_path: Path
) -> None:
    """Sanity check that these two real files are actually meant to be
    merged together: every base-file رقم الحيازة should have a match in
    the external file, once leading zeros are normalized."""
    from src.domain.value_objects.holding_id import normalize_for_join

    base_ids = {normalize_for_join(r.holding_id_raw) for r in _read_system_records(base_file_path)}
    external_ids = {
        normalize_for_join(r.holding_id_raw) for r in _read_external_records(secondary_file_path)
    }
    assert base_ids <= external_ids


# --- Seasonal-survey mapping (real file, older secondary-file format) ------


def test_seasonal_mapping_excludes_laghi_rows(seasonal_secondary_file_path: Path) -> None:
    excluded_pairs: set[tuple[str, str | None]] = set()
    kept_pairs: set[tuple[str, str | None]] = set()

    for cells in _seasonal_raw_rows(seasonal_secondary_file_path):
        holding_id = str(cells["K"]).strip() if cells["K"] is not None else None
        if not holding_id:
            continue
        name = str(cells["N"]).strip() if cells["N"] is not None else None
        pair = (holding_id, name)
        (excluded_pairs if _is_excluded_row(cells) else kept_pairs).add(pair)

    purely_excluded_pairs = excluded_pairs - kept_pairs
    assert purely_excluded_pairs, "sanity check: expected a لاغى row with no non-excluded duplicate"

    records = _read_seasonal_records(seasonal_secondary_file_path)
    present_pairs = {(r.holding_id_raw, r.holder_name) for r in records}
    assert purely_excluded_pairs.isdisjoint(present_pairs)


def test_seasonal_mapping_reads_valid_national_id(seasonal_secondary_file_path: Path) -> None:
    expected = None
    for cells in _seasonal_raw_rows(seasonal_secondary_file_path):
        if _is_excluded_row(cells):
            continue
        if cells["AD"] is not None and cells["K"] is not None:
            raw_id = str(cells["AD"]).strip()
            expected_id = raw_id.zfill(14) if raw_id.isdigit() else raw_id
            expected = (str(cells["K"]).strip(), expected_id)
            break
    assert expected is not None, "sanity check: expected a non-excluded row with a national ID"

    records = _read_seasonal_records(seasonal_secondary_file_path)
    assert any(
        (r.holding_id_raw, r.national_id) == expected and len(r.national_id or "") == 14
        for r in records
    )


def test_seasonal_mapping_basin_numeric_value_maps_to_basin_code(
    seasonal_secondary_file_path: Path,
) -> None:
    expected = None
    for cells in _seasonal_raw_rows(seasonal_secondary_file_path):
        if _is_excluded_row(cells):
            continue
        af = cells["AF"]
        if isinstance(af, int | float) and cells["K"] is not None:
            expected = (str(cells["K"]).strip(), str(af).strip())
            break
    assert expected is not None, "sanity check: expected a non-excluded row with a numeric الجوض"

    records = _read_seasonal_records(seasonal_secondary_file_path)
    assert any(
        r.holding_id_raw == expected[0] and r.basin_code == expected[1] and r.basin_name is None
        for r in records
    )


def test_seasonal_mapping_basin_text_value_maps_to_basin_name(
    seasonal_secondary_file_path: Path,
) -> None:
    expected = None
    for cells in _seasonal_raw_rows(seasonal_secondary_file_path):
        if _is_excluded_row(cells):
            continue
        af = cells["AF"]
        if isinstance(af, str) and af.strip() and cells["K"] is not None:
            expected = (str(cells["K"]).strip(), af.strip())
            break
    assert expected is not None, "sanity check: expected a non-excluded row with a text الجوض"

    records = _read_seasonal_records(seasonal_secondary_file_path)
    assert any(
        r.holding_id_raw == expected[0] and r.basin_name == expected[1] and r.basin_code is None
        for r in records
    )


def test_seasonal_mapping_area_uses_total_group(seasonal_secondary_file_path: Path) -> None:
    expected = None
    for cells in _seasonal_raw_rows(seasonal_secondary_file_path):
        if _is_excluded_row(cells):
            continue
        if cells["AA"] is not None and cells["K"] is not None:
            expected = (str(cells["K"]).strip(), cells["AA"], cells["AB"], cells["AC"])
            break
    assert expected is not None, "sanity check: expected a non-excluded row with جملة area values"

    holding_id, sahm, qirat, feddan = expected
    records = _read_seasonal_records(seasonal_secondary_file_path)
    assert any(
        r.holding_id_raw == holding_id
        and r.sahm == sahm
        and r.qirat == qirat
        and r.feddan == feddan
        for r in records
    )


def test_seasonal_mapping_borders_and_land_number_stay_none(
    seasonal_secondary_file_path: Path,
) -> None:
    """The seasonal mapping declares no border/land-number fields, so
    MappedFileReader must leave them None — never invented."""
    records = _read_seasonal_records(seasonal_secondary_file_path)
    assert all(r.east is None for r in records)
    assert all(r.land_number is None for r in records)


# --- Exclusion toggle (synthetic, fast) -------------------------------------

_TOGGLE_CONFIG: dict[str, Any] = {
    "sheet_name": "Sheet1",
    "data_starts_at_row": 2,
    "join_key_column": "A",
    "exclude_when": {"column": "B", "values": ["لاغى", "لاغي"]},
    "fields": {"اسم_الحائز": "C"},
}


def _write_toggle_workbook(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    # openpyxl's append() always writes starting at column A, regardless
    # of what the header text says — these headers land in A/B/C, which
    # is what _TOGGLE_CONFIG's column letters must match.
    sheet.append(["ID", "Excluded Flag", "Holder Name"])
    sheet.append(["1", "لاغى", "Excluded Holder"])
    sheet.append(["2", None, "Kept Holder"])
    workbook.save(path)


def test_excludes_laghi_rows_by_default(tmp_path: Path) -> None:
    path = tmp_path / "secondary.xlsx"
    _write_toggle_workbook(path)
    mapper = YamlColumnMapper(_TOGGLE_CONFIG["fields"])

    records = MappedFileReader(path, mapper, _TOGGLE_CONFIG).read()

    assert {r.holding_id_raw for r in records} == {"2"}


def test_include_laghi_rows_when_apply_exclusion_is_false(tmp_path: Path) -> None:
    path = tmp_path / "secondary.xlsx"
    _write_toggle_workbook(path)
    mapper = YamlColumnMapper(_TOGGLE_CONFIG["fields"])

    records = MappedFileReader(path, mapper, _TOGGLE_CONFIG, apply_exclusion=False).read()

    assert {r.holding_id_raw for r in records} == {"1", "2"}


# --- New Iteration 3 behavior: join key independent of field names ---------


def test_join_key_extracted_from_join_key_column_not_a_named_field(tmp_path: Path) -> None:
    """The join key must come from `join_key_column`, not a field named
    رقم_الحيازة — proving the two are now independent, per Task A."""
    path = tmp_path / "custom.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Some ID Column", "Holder Name"])
    sheet.append(["XYZ-123", "Someone"])
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",  # deliberately not named رقم_الحيازة in fields
        "fields": {"اسم_الحائز": "B"},
    }
    mapper = YamlColumnMapper(config["fields"])

    records = MappedFileReader(path, mapper, config).read()

    assert len(records) == 1
    assert records[0].holding_id_raw == "XYZ-123"
    assert records[0].holder_name == "Someone"


def test_missing_declared_column_data_is_handled_gracefully(tmp_path: Path) -> None:
    """A field the mapping declares but which is blank in a given row
    must resolve to None, not raise."""
    path = tmp_path / "sparse.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ID", "Name", "National ID"])
    sheet.append(["1", "Someone", None])
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"اسم_الحائز": "B", "الرقم_القومي": "C"},
    }
    mapper = YamlColumnMapper(config["fields"])

    records = MappedFileReader(path, mapper, config).read()

    assert len(records) == 1
    assert records[0].national_id is None


def test_row_with_blank_join_key_is_skipped_not_errored(tmp_path: Path) -> None:
    path = tmp_path / "with_blank.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ID", "Name"])
    sheet.append([None, "No ID"])
    sheet.append(["1", "Has ID"])
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"اسم_الحائز": "B"},
    }
    mapper = YamlColumnMapper(config["fields"])

    records = MappedFileReader(path, mapper, config).read()

    assert len(records) == 1
    assert records[0].holder_name == "Has ID"


# --- Sanity: both real mapping files still load cleanly ---------------------


def test_all_default_mappings_load_without_error() -> None:
    system_config = load_mapping(SYSTEM_MAPPING_PATH)
    external_config = load_mapping(EXTERNAL_MAPPING_PATH)
    seasonal_config = load_mapping(SEASONAL_MAPPING_PATH)

    assert "join_key_column" in system_config
    assert "join_key_column" in external_config
    assert "join_key_column" in seasonal_config
    assert system_config["join_key_column"] == "N"
    assert external_config["join_key_column"] == "R"
    assert seasonal_config["join_key_column"] == "K"


# --- Sheet-name drift between exports of the same report ---------------------
#
# Report exports of the same underlying report type sometimes get a
# different auto-generated sheet name per export run (observed in
# practice with the system-file report). Since these reports are
# always single-sheet, the reader falls back to the workbook's only
# sheet rather than failing the whole run.


def _write_single_sheet_workbook(path: Path, sheet_title: str) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.append(["ID", "Name"])
    sheet.append(["1", "Someone"])
    workbook.save(path)


def test_falls_back_to_the_only_sheet_when_configured_name_is_absent(tmp_path: Path) -> None:
    path = tmp_path / "renamed_sheet.xlsx"
    _write_single_sheet_workbook(path, sheet_title="SomeOtherAutoGeneratedName")

    config: dict[str, Any] = {
        "sheet_name": "ExpectedSheetName",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"اسم_الحائز": "B"},
    }
    mapper = YamlColumnMapper(config["fields"])
    reader = MappedFileReader(path, mapper, config)

    records = reader.read()

    assert len(records) == 1
    assert records[0].holder_name == "Someone"
    assert reader.sheet_fallback_warning is not None
    assert "ExpectedSheetName" in reader.sheet_fallback_warning
    assert "SomeOtherAutoGeneratedName" in reader.sheet_fallback_warning


def test_sheet_fallback_warning_is_none_when_configured_name_matches(tmp_path: Path) -> None:
    path = tmp_path / "matching_sheet.xlsx"
    _write_single_sheet_workbook(path, sheet_title="Sheet1")

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"اسم_الحائز": "B"},
    }
    mapper = YamlColumnMapper(config["fields"])
    reader = MappedFileReader(path, mapper, config)

    reader.read()

    assert reader.sheet_fallback_warning is None


def test_raises_when_sheet_name_is_absent_and_workbook_has_multiple_sheets(
    tmp_path: Path,
) -> None:
    path = tmp_path / "multi_sheet.xlsx"
    workbook = openpyxl.Workbook()
    first_sheet = workbook.active
    assert first_sheet is not None
    first_sheet.title = "FirstSheet"
    workbook.create_sheet("SecondSheet")
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "ExpectedSheetName",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"اسم_الحائز": "B"},
    }
    mapper = YamlColumnMapper(config["fields"])

    with pytest.raises(KeyError, match="ExpectedSheetName"):
        MappedFileReader(path, mapper, config).read()


# --- read_first_field_value --------------------------------------------------


def test_read_first_field_value_returns_first_non_empty_value(tmp_path: Path) -> None:
    path = tmp_path / "society.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ID", "Name", "Society"])
    sheet.append(["1", "Someone", "ميت فضالة-الائتمان الزراعي"])
    sheet.append(["2", "Someone Else", "ميت فضالة-الائتمان الزراعي"])
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"اسم_الحائز": "B", "الجمعيه": "C"},
    }

    value = read_first_field_value(path, config, "الجمعيه")

    assert value == "ميت فضالة-الائتمان الزراعي"


def test_read_first_field_value_skips_blank_rows(tmp_path: Path) -> None:
    path = tmp_path / "society_blank_first.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ID", "Society"])
    sheet.append(["1", None])
    sheet.append(["2", "ميت فضالة-الائتمان الزراعي"])
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"الجمعيه": "B"},
    }

    value = read_first_field_value(path, config, "الجمعيه")

    assert value == "ميت فضالة-الائتمان الزراعي"


def test_read_first_field_value_returns_none_when_field_not_mapped(tmp_path: Path) -> None:
    path = tmp_path / "no_society.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ID"])
    sheet.append(["1"])
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {},
    }

    assert read_first_field_value(path, config, "الجمعيه") is None


def test_read_first_field_value_returns_none_when_all_scanned_rows_blank(tmp_path: Path) -> None:
    path = tmp_path / "all_blank_society.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ID", "Society"])
    sheet.append(["1", None])
    workbook.save(path)

    config: dict[str, Any] = {
        "sheet_name": "Sheet1",
        "data_starts_at_row": 2,
        "join_key_column": "A",
        "fields": {"الجمعيه": "B"},
    }

    assert read_first_field_value(path, config, "الجمعيه") is None
